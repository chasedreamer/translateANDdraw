from openpyxl import Workbook
from openpyxl import load_workbook

import time
import urllib
import hashlib
import requests
import json

class Geoconv(object):
    my_ak = '------------------------'
    my_sk = 'fFyAglwnmgxnYI6c---------------------------'

    def __init__(self):
        pass

    # GPS坐标转换为百度坐标
    #解释：http://lbsyun.baidu.com/index.php?title=webapi/guide/changeposition
    #API：http://api.map.baidu.com/geoconv/v1/?coords=114.21892734521,29.575429778924&from=1&to=5&ak=你的密钥 //GET
    def wgs84tobd09(self,lon,lat):
        queryStr = '/geoconv/v1/?coords={},{}&from=1&to=5&ak={}'.format(lon,lat,self.my_ak)
        # print queryStr
        # 对queryStr进行转码，safe内的保留字符不转换
        encodedStr = urllib.parse.quote(queryStr, safe="/:=&?#+!$,;'@()*[]")

        # 在最后直接追加上yoursk
        rawStr = encodedStr + self.my_sk

        # md5计算出的sn值
        my_sn = '----------------------'  ###hashlib.md5(urllib.parse.quote_plus(rawStr)).hexdigest()
        # print my_sn
        url = 'http://api.map.baidu.com' + queryStr + "&sn=" + my_sn
        # print url

        res = requests.get(url)
        if res.status_code != requests.codes.ok:
            print('404')
            res = requests.get(url)
        # print '*' * 10
        # get收到的内容
        json_str = res.content
        #print(json_str)
        dictData = json.loads(json_str)
        # print dictData["result"][0]["x"]
        # print dictData["result"][0]["y"]
        if dictData["status"] != 0:
            print("status=",dictData["status"] )
            return 0,0
        return dictData["result"][0]["x"],dictData["result"][0]["y"]

# 默认可读写，若有需要可以指定write_only和read_only为True
wb_read = load_workbook('1.xlsx')
a_sheet = wb_read.get_sheet_by_name('Sheet1')
# 获得sheet名
print(a_sheet.title)

wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
#ws['A1'] = 42

# Rows can also be appended
#ws.append([1, 2, 3])

# Python types will automatically be converted
#import datetime
#ws['A2'] = datetime.datetime.now()


print(a_sheet.max_row)
print(a_sheet.max_column)
myconv = Geoconv()
for i in range(2, a_sheet.max_row):
    #for j in range(1, a_sheet.max_column):
    #print(a_sheet.cell(row=i, column=1).value,a_sheet.cell(row=i, column=2).value)
    lon = int(a_sheet.cell(row=i, column=1).value)
    lat = int(a_sheet.cell(row=i, column=2).value)
    print(i,'/', a_sheet.max_row)
    ws.cell(row=i, column=1).value,ws.cell(row=i, column=2).value =myconv.wgs84tobd09(lon*0.000001,lat*0.000001) 
    if i%1000 ==0:
        wb.save("sample.xlsx")
     
	
# Save the file
wb.save("sample.xlsx")	
	